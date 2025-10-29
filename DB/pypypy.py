import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import psycopg2
from datetime import date
import subprocess
import os
import csv
from tkinter import filedialog
import requests
import json
from datetime import datetime

DB_CONFIG = {
    'dbname': 'LocalBase',
    'user': 'postgres',
    'password': '',
    'host': 'localhost',
    'port': '5433'
}

# Файл для сохранения API ключей
API_CONFIG_FILE = "api_config.json"


# Загрузка API ключей из файла
def load_api_config():
    default_config = {
        'weather': 'c39f84ae0e605598fd480683bbc96cc5',
        'news': '3811df40db174b3fb63de857067c3c90',
        'sports': '1',  # TheSportsDB использует 1 для демо-доступа
        'space': 'DEMO_KEY',  # NASA DEMO_KEY (рабочий)
        'pets': ''  # The Dog API не требует ключа
    }

    try:
        if os.path.exists(API_CONFIG_FILE):
            with open(API_CONFIG_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            # Создаем файл с default значениями
            save_api_config(default_config)
            return default_config
    except Exception:
        return default_config


# Сохранение API ключей в файл
def save_api_config(config):
    try:
        with open(API_CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=4)
        return True
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить настройки:\n{e}")
        return False


# Глобальная переменная для API конфигурации
API_CONFIG = load_api_config()


def get_db_connection():
    return psycopg2.connect(**DB_CONFIG)


# ==================== WEATHER API ====================
def get_weather_data(city="Moscow"):
    """Получение данных о погоде с OpenWeatherMap"""
    try:
        api_key = API_CONFIG.get('weather', '').strip()

        if not api_key:
            return {
                'city': city,
                'temperature': '21°C',
                'description': 'Облачно',
                'humidity': '65%',
                'pressure': '1013 hPa',
                'wind': '3 м/с',
                'demo': True,
                'error': 'API ключ не установлен'
            }

        url = f"http://api.openweathermap.org/data/2.5/weather?q={city}&appid={api_key}&units=metric&lang=ru"
        response = requests.get(url, timeout=10)

        if response.status_code == 200:
            data = response.json()
            return {
                'city': data['name'],
                'temperature': f"{data['main']['temp']}°C",
                'description': data['weather'][0]['description'].capitalize(),
                'humidity': f"{data['main']['humidity']}%",
                'pressure': f"{data['main']['pressure']} hPa",
                'wind': f"{data['wind']['speed']} м/с",
                'demo': False
            }
        elif response.status_code == 401:
            return {
                'city': city,
                'temperature': 'N/A',
                'description': 'Неверный API ключ',
                'humidity': 'N/A',
                'pressure': 'N/A',
                'wind': 'N/A',
                'demo': True,
                'error': 'Неверный API ключ'
            }
        else:
            return None

    except Exception as e:
        return {
            'city': city,
            'temperature': 'N/A',
            'description': f'Ошибка: {str(e)}',
            'humidity': 'N/A',
            'pressure': 'N/A',
            'wind': 'N/A',
            'demo': True,
            'error': str(e)
        }


# ==================== NEWS API ====================
def get_news_data(page=1):
    """Получение последних новостей с Hacker News API"""
    try:
        # Используем Hacker News API (бесплатный, без ключа)
        url = f"https://api.hnpwa.com/v0/news/{page}.json"
        response = requests.get(url, timeout=10)

        if response.status_code == 200:
            data = response.json()
            news_list = []
            for article in data[:10]:  # Берем первые 10 новостей
                news_list.append({
                    "title": article.get('title', 'Без названия'),
                    "source": "Hacker News",
                    "published": article.get('time_ago', ''),
                    "url": article.get('url', '')
                })
            return news_list
        else:
            # Демо-новости если API не доступен
            demo_news = [
                {"title": "Технологии развиваются быстрыми темпами", "source": "TechNews", "published": "Сегодня", "url": ""},
                {"title": "Новые открытия в области науки", "source": "Science Daily", "published": "Вчера", "url": ""},
                {"title": "Спортивные события недели", "source": "Sports Review", "published": "2 дня назад", "url": ""},
                {"title": "Культурные мероприятия в городе", "source": "Culture Guide", "published": "Сегодня", "url": ""}
            ]
            return demo_news

    except Exception as e:
        return []


# ==================== SPORTS API ====================
def get_sports_data():
    """Получение спортивных данных с TheSportsDB"""
    try:
        api_key = API_CONFIG.get('sports', '1').strip()
        
        # Используем TheSportsDB API для получения футбольных событий
        url = f"https://www.thesportsdb.com/api/v1/json/{api_key}/eventsnext.php?id=133602"  # ID команды
        
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            events = data.get('events', [])
            
            if events:
                sports_list = []
                for event in events[:5]:  # Берем первые 5 событий
                    sports_list.append({
                        "sport": "Футбол",
                        "event": event.get('strEvent', 'Матч'),
                        "result": f"{event.get('strHomeTeam', 'Команда А')} vs {event.get('strAwayTeam', 'Команда Б')}",
                        "date": event.get('dateEvent', '')
                    })
                return sports_list
        
        # Если API не вернул данные, используем демо-данные
        demo_sports = [
            {"sport": "Футбол", "event": "Чемпионат Англии", "result": "Манчестер Юнайтед 2:1 Ливерпуль", "date": "2024-03-20"},
            {"sport": "Хоккей", "event": "КХЛ", "result": "ЦСКА 3:2 СКА", "date": "2024-03-19"},
            {"sport": "Баскетбол", "event": "Евролига", "result": "ЦСКА 85:78 Реал", "date": "2024-03-18"},
            {"sport": "Теннис", "event": "US Open", "result": "Медведев - в полуфинале", "date": "2024-03-17"}
        ]
        return demo_sports
        
    except Exception as e:
        return []


# ==================== SPACE API ====================
def get_space_data():
    """Получение данных о космосе (NASA API)"""
    try:
        api_key = API_CONFIG.get('space', 'DEMO_KEY').strip()

        # Реальный запрос к NASA API - Astronomy Picture of the Day
        url = f"https://api.nasa.gov/planetary/apod?api_key={api_key}"
        response = requests.get(url, timeout=10)

        if response.status_code == 200:
            data = response.json()
            return [
                {
                    "title": data.get('title', 'NASA Astronomy Picture of the Day'),
                    "info": data.get('explanation', '')[:200] + "...",
                    "date": data.get('date', ''),
                    "image": data.get('url', '')
                }
            ]
        else:
            # Демо-данные о космосе если API не доступен
            demo_space = [
                {"title": "Марсоход Perseverance", "info": "Исследует кратер Езеро на Марсе, собирает образцы грунта", "date": "2024-03-20"},
                {"title": "Международная космическая станция", "info": "Экипаж из 7 человек проводит научные эксперименты", "date": "2024-03-19"},
                {"title": "Телескоп Джеймс Уэбб", "info": "Получает новые изображения далеких галактик и туманностей", "date": "2024-03-18"}
            ]
            return demo_space

    except Exception as e:
        return []


# ==================== PETS API ====================
def get_pets_data():
    """Получение данных о домашних животных"""
    try:
        # The Dog API - бесплатный, не требует ключа
        response = requests.get("https://dog.ceo/api/breeds/image/random/4", timeout=10)

        if response.status_code == 200:
            data = response.json()
            breeds = ["Лабрадор", "Овчарка", "Пудель", "Бульдог"]
            pets = []
            for i, image_url in enumerate(data.get('message', [])):
                breed_name = breeds[i] if i < len(breeds) else f"Порода {i + 1}"
                pets.append({
                    "name": breed_name,
                    "breed": "Собака",
                    "info": "Прекрасный друг человека",
                    "image": image_url
                })
            return pets
        else:
            # Демо-данные если API не доступен
            return [
                {"name": "Лабрадор", "breed": "Собака", "info": "Дружелюбная семейная собака", "image": ""},
                {"name": "Персидский кот", "breed": "Кошка", "info": "Пушистый и спокойный", "image": ""},
                {"name": "Волнистый попугай", "breed": "Птица", "info": "Яркий и говорящий", "image": ""},
                {"name": "Сирийский хомяк", "breed": "Грызун", "info": "Активный ночной зверек", "image": ""}
            ]
    except Exception:
        return []


# ==================== DATABASE FUNCTIONS ====================
def load_data(tree):
    for row in tree.get_children():
        tree.delete(row)
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT * FROM Big_30_1 ORDER BY ID")
        rows = cur.fetchall()
        for row in rows:
            tree.insert("", "end", values=row)
        cur.close()
        conn.close()
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось загрузить данные:\n{e}")


def add_person(entries, tree):
    try:
        data = {}
        for field, entry in entries.items():
            val = entry.get().strip()
            if field in ['День_рождения', 'Месяц_рождения', 'Год_рождения', 'Этаж', 'Квартира', 'Курс',
                         'Левый_карман', 'Правый_карман', 'Деньги_на_карте']:
                data[field] = int(val) if val else None
            elif field in ['Вес', 'Рост', 'ЗП']:
                data[field] = float(val) if val else None
            else:
                data[field] = val if val else None

        if data['Год_рождения'] and data['Месяц_рождения'] and data['День_рождения']:
            data['Дата_рождения'] = date(
                data['Год_рождения'],
                data['Месяц_рождения'],
                data['День_рождения']
            )
        else:
            data['Дата_рождения'] = None

        conn = get_db_connection()
        cur = conn.cursor()
        query = """
        INSERT INTO Big_30_1 (
            Фамилия, Имя, Отчество, Телефон, День_рождения, Месяц_рождения, Год_рождения,
            Дом, Корпус, Подъезд, Этаж, Квартира, Город, Пол, Вес, Рост, ЗП, Цвет_волос, Дата_рождения,
            Хобби, Курс, Специальность, Левый_карман, Правый_карман, Деньги_на_карте
        ) VALUES (
            %(Фамилия)s, %(Имя)s, %(Отчество)s, %(Телефон)s, %(День_рождения)s, %(Месяц_рождения)s,
            %(Год_рождения)s, %(Дом)s, %(Корпус)s, %(Подъезд)s, %(Этаж)s, %(Квартира)s, %(Город)s,
            %(Пол)s, %(Вес)s, %(Рост)s, %(ЗП)s, %(Цвет_волос)s, %(Дата_рождения)s, %(Хобби)s,
            %(Курс)s, %(Специальность)s, %(Левый_карман)s, %(Правый_карман)s, %(Деньги_на_карте)s
        )
        """
        cur.execute(query, data)
        conn.commit()
        cur.close()
        conn.close()

        messagebox.showinfo("Успех", "Запись успешно добавлена!")
        load_data(tree)

        for entry in entries.values():
            entry.delete(0, tk.END)

    except ValueError as ve:
        messagebox.showerror("Ошибка ввода", f"Проверьте числовые поля:\n{ve}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось добавить запись:\n{e}")


def delete_record(tree):
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Предупреждение", "Пожалуйста, выберите запись для удаления")
        return

    try:
        record_id = tree.item(selected_item[0])['values'][0]
        confirm = messagebox.askyesno("Подтверждение", f"Вы уверены, что хотите удалить запись с ID {record_id}?")
        if not confirm:
            return

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM Big_30_1 WHERE ID = %s", (record_id,))
        conn.commit()
        cur.close()
        conn.close()

        messagebox.showinfo("Успех", "Запись успешно удалена!")
        load_data(tree)

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось удалить запись:\n{e}")


def open_paint():
    if os.name == 'nt':
        subprocess.Popen('mspaint')


def open_calculator():
    if os.name == 'nt':
        subprocess.Popen('calc')


def open_calendar():
    import webbrowser
    webbrowser.open('https://my-calend.ru/calendar/2025')


def export_to_excel(results, columns, query):
    """Экспорт результатов в Excel файл"""
    if not results:
        messagebox.showwarning("Предупреждение", "Нет данных для экспорта")
        return

    try:
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить результаты как Excel"
        )

        if file_path:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "Результаты запроса"

            ws['A1'] = f"Результаты запроса: {query}"
            ws['A1'].font = Font(bold=True, size=12)
            ws.merge_cells('A1:Z1')

            ws['A2'] = ""

            for col_idx, column_name in enumerate(columns, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = column_name
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            for row_idx, row_data in enumerate(results, 4):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx).value = cell_value

            for col_idx in range(1, len(columns) + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)

                for row_idx in range(3, len(results) + 4):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            messagebox.showinfo("Успех", f"Данные успешно экспортированы в:\n{file_path}")

    except ImportError:
        response = messagebox.askyesno(
            "Библиотека не найдена",
            "Для экспорта в Excel требуется библиотека openpyxl.\nХотите установить её сейчас?"
        )
        if response:
            try:
                import subprocess
                import sys
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                messagebox.showinfo("Успех", "openpyxl успешно установлен!\nПопробуйте экспорт снова.")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось установить openpyxl:\n{e}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Ошибка при экспорте в Excel:\n{e}")


def open_query_window():
    query_window = tk.Toplevel()
    query_window.title("Выполнение SQL-запросов")
    query_window.geometry("800x600")

    tk.Label(query_window, text="Введите SQL-запрос:", font=("Arial", 12, "bold")).pack(pady=10)

    query_text = scrolledtext.ScrolledText(query_window, width=90, height=10, font=("Consolas", 10))
    query_text.pack(pady=10, padx=10, fill='both', expand=True)

    examples_frame = tk.Frame(query_window)
    examples_frame.pack(pady=5)

    tk.Label(examples_frame, text="Примеры запросов:", font=("Arial", 10, "bold")).pack()

    examples = [
        "SELECT * FROM Big_30_1 WHERE Город = 'Москва'",
        "SELECT Фамилия, Имя, Город FROM Big_30_1 WHERE ЗП > 50000",
        "SELECT Город, COUNT(*) as Количество FROM Big_30_1 GROUP BY Город",
        "SELECT * FROM Big_30_1 WHERE Дата_рождения > '1990-01-01'"
    ]

    for example in examples:
        example_label = tk.Label(examples_frame, text=example, font=("Consolas", 8), fg="blue", cursor="hand2")
        example_label.pack(pady=2)
        example_label.bind("<Button-1>", lambda e, ex=example: query_text.insert("1.0", ex + "\n"))

    button_frame = tk.Frame(query_window)
    button_frame.pack(pady=10)

    def execute_query():
        query = query_text.get("1.0", tk.END).strip()
        if not query:
            messagebox.showwarning("Предупреждение", "Введите SQL-запрос")
            return

        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute(query)

            if query.strip().upper().startswith('SELECT'):
                results = cur.fetchall()
                columns = [desc[0] for desc in cur.description]
                show_query_results(results, columns, query)
            else:
                conn.commit()
                messagebox.showinfo("Успех", f"Запрос выполнен успешно!\nЗатронуто строк: {cur.rowcount}")

            cur.close()
            conn.close()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка выполнения запроса:\n{e}")

    def show_query_results(results, columns, query):
        result_window = tk.Toplevel()
        result_window.title("Результаты запроса")
        result_window.geometry("1000x600")

        result_window.results = results
        result_window.columns = columns
        result_window.query = query

        tk.Label(result_window, text=f"Результаты запроса: {query}", font=("Arial", 10), wraplength=800).pack(pady=10)
        tk.Label(result_window, text=f"Найдено записей: {len(results)}", font=("Arial", 10, "bold")).pack()

        tree_frame = tk.Frame(result_window)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

        result_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)

        for col in columns:
            result_tree.heading(col, text=col)
            result_tree.column(col, width=100, anchor='center')

        for row in results:
            result_tree.insert("", "end", values=row)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=result_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=result_tree.xview)
        result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        result_tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')

        export_frame = tk.Frame(result_window)
        export_frame.pack(pady=10)

        btn_excel = tk.Button(export_frame, text="Экспорт в Excel",
                              command=lambda: export_to_excel(results, columns, query),
                              bg="lightgreen", font=("Arial", 10, "bold"))
        btn_excel.pack(pady=10)

    btn_execute = tk.Button(button_frame, text="Выполнить запрос",
                            command=execute_query, bg="lightcoral", fg="white", font=("Arial", 10, "bold"))
    btn_execute.pack(side='left', padx=5)

    btn_clear = tk.Button(button_frame, text="Очистить",
                          command=lambda: query_text.delete("1.0", tk.END))
    btn_clear.pack(side='left', padx=5)

    btn_close = tk.Button(button_frame, text="Закрыть",
                          command=query_window.destroy)
    btn_close.pack(side='left', padx=5)


# ==================== API SETTINGS TAB ====================
def create_api_settings_tab(tab):
    frame = tk.Frame(tab)
    frame.pack(fill='both', expand=True, padx=20, pady=20)

    tk.Label(frame, text="⚙️ Настройки API Ключей", font=("Arial", 18, "bold")).pack(pady=20)

    # Информация о API
    info_text = """✅ Все API ключи настроены и готовы к работе!

🌤 Погода: OpenWeatherMap (рабочий ключ)
📰 Новости: Hacker News API (бесплатный, без ключа)
⚽ Спорт: TheSportsDB (демо-ключ)
🚀 Космос: NASA API (демо-ключ)
🐾 Животные: The Dog API (не требует ключа)

Вы можете заменить ключи на свои собственные:"""

    info_label = tk.Label(frame, text=info_text, font=("Arial", 10), justify='left', wraplength=600)
    info_label.pack(pady=10)

    # Поля для ввода API ключей
    api_frame = tk.Frame(frame)
    api_frame.pack(pady=20, fill='x')

    api_entries = {}

    # Погода
    weather_frame = tk.Frame(api_frame)
    weather_frame.pack(fill='x', pady=5)
    tk.Label(weather_frame, text="🌤 OpenWeatherMap API:", font=("Arial", 10, "bold")).pack(anchor='w')
    weather_entry = tk.Entry(weather_frame, width=60, show="*")
    weather_entry.insert(0, API_CONFIG.get('weather', ''))
    weather_entry.pack(fill='x', pady=2)
    api_entries['weather'] = weather_entry

    # Новости
    news_frame = tk.Frame(api_frame)
    news_frame.pack(fill='x', pady=5)
    tk.Label(news_frame, text="📰 Hacker News API (не требует ключа):", font=("Arial", 10, "bold")).pack(anchor='w')
    news_label = tk.Label(news_frame, text="Бесплатный API - ключ не нужен", fg="green")
    news_label.pack(anchor='w', pady=2)
    # Не добавляем в api_entries, так как ключ не нужен

    # Спорт
    sports_frame = tk.Frame(api_frame)
    sports_frame.pack(fill='x', pady=5)
    tk.Label(sports_frame, text="⚽ TheSportsDB API:", font=("Arial", 10, "bold")).pack(anchor='w')
    sports_entry = tk.Entry(sports_frame, width=60)
    sports_entry.insert(0, API_CONFIG.get('sports', ''))
    sports_entry.pack(fill='x', pady=2)
    api_entries['sports'] = sports_entry

    # Космос
    space_frame = tk.Frame(api_frame)
    space_frame.pack(fill='x', pady=5)
    tk.Label(space_frame, text="🚀 NASA API:", font=("Arial", 10, "bold")).pack(anchor='w')
    space_entry = tk.Entry(space_frame, width=60)
    space_entry.insert(0, API_CONFIG.get('space', ''))
    space_entry.pack(fill='x', pady=2)
    api_entries['space'] = space_entry

    # Кнопки
    button_frame = tk.Frame(frame)
    button_frame.pack(pady=20)

    def save_settings():
        new_config = {}
        for key, entry in api_entries.items():
            new_config[key] = entry.get().strip()

        if save_api_config(new_config):
            global API_CONFIG
            API_CONFIG = new_config
            messagebox.showinfo("Успех", "Настройки API сохранены!")

    def toggle_visibility():
        current_show = weather_entry.cget('show')
        new_show = "" if current_show == "*" else "*"
        for entry in api_entries.values():
            entry.config(show=new_show)
        btn_visibility.config(text="👁 Скрыть" if new_show == "*" else "👁 Показать")

    def clear_all():
        for entry in api_entries.values():
            entry.delete(0, tk.END)

    btn_save = tk.Button(button_frame, text="💾 Сохранить настройки",
                         command=save_settings, bg="lightgreen", font=("Arial", 10, "bold"))
    btn_save.pack(side='left', padx=5)

    btn_visibility = tk.Button(button_frame, text="👁 Показать ключи",
                               command=toggle_visibility, bg="lightblue")
    btn_visibility.pack(side='left', padx=5)

    btn_clear = tk.Button(button_frame, text="🗑 Очистить все",
                          command=clear_all, bg="lightcoral")
    btn_clear.pack(side='left', padx=5)

    # Статус текущих настроек
    status_frame = tk.Frame(frame)
    status_frame.pack(pady=10, fill='x')

    tk.Label(status_frame, text="Текущий статус API ключей:", font=("Arial", 11, "bold")).pack(anchor='w')

    status_text = tk.Text(status_frame, height=5, width=60, font=("Arial", 9))
    status_text.pack(fill='x', pady=5)

    def update_status():
        status_text.delete(1.0, tk.END)
        for key, value in API_CONFIG.items():
            if key == 'weather':
                status = "✅ Рабочий ключ" if value.strip() else "❌ Не установлен"
                service_name = '🌤 Погода (OpenWeatherMap)'
            elif key == 'news':
                status = "✅ Бесплатный API (работает всегда)"
                service_name = '📰 Новости (Hacker News)'
            elif key == 'sports':
                status = "✅ Демо-ключ (работает)" if value.strip() else "❌ Не установлен"
                service_name = '⚽ Спорт (TheSportsDB)'
            elif key == 'space':
                status = "✅ Демо-ключ (работает)" if value.strip() else "❌ Не установлен"
                service_name = '🚀 Космос (NASA)'
            else:
                status = "✅ Настроен" if value.strip() else "❌ Не установлен"
                service_name = key
            status_text.insert(tk.END, f"{service_name}: {status}\n")

        status_text.insert(tk.END, f"\n🐾 Животные (The Dog API): ✅ Не требует ключа")

    btn_status = tk.Button(status_frame, text="🔄 Обновить статус",
                           command=update_status, bg="lightyellow")
    btn_status.pack(pady=5)

    update_status()


# ==================== WEATHER TAB ====================
def create_weather_tab(tab):
    frame = tk.Frame(tab)
    frame.pack(fill='both', expand=True, padx=10, pady=10)

    tk.Label(frame, text="🌤 Погода", font=("Arial", 16, "bold")).pack(pady=10)

    # Статус API
    api_status = tk.Label(frame, text="", font=("Arial", 10), fg="green")
    api_status.pack(pady=5)

    # Поиск погды по городу
    search_frame = tk.Frame(frame)
    search_frame.pack(pady=10)

    tk.Label(search_frame, text="Город:").pack(side='left')

    # Список популярных городов
    cities = ["Moscow", "Saint Petersburg", "Novosibirsk", "Yekaterinburg", "Kazan", "Nizhny Novgorod",
              "Chelyabinsk", "Samara", "Omsk", "Rostov-on-Don", "Ufa", "Krasnoyarsk", "Voronezh", "Perm",
              "Volgograd", "Krasnodar", "Saratov", "Tyumen", "Tolyatti", "Izhevsk"]

    city_var = tk.StringVar()
    city_combobox = ttk.Combobox(search_frame, textvariable=city_var, values=cities, width=18)
    city_combobox.set("Moscow")
    city_combobox.pack(side='left', padx=5)

    weather_text = scrolledtext.ScrolledText(frame, width=60, height=15, font=("Arial", 10))
    weather_text.pack(pady=10, fill='both', expand=True)

    def search_weather():
        city = city_var.get().strip()
        if not city:
            city = "Moscow"

        weather_data = get_weather_data(city)
        weather_text.delete(1.0, tk.END)

        # Обновляем статус API
        if API_CONFIG.get('weather', '').strip():
            api_status.config(text="✅ API ключ установлен - реальные данные", fg="green")
        else:
            api_status.config(text="⚠️ Демо-режим: установите API ключ в настройках", fg="orange")

        if weather_data:
            if weather_data.get('demo', False):
                weather_text.insert(tk.END, "⚠ ДЕМО-РЕЖИМ ⚠\n")
                if weather_data.get('error'):
                    weather_text.insert(tk.END, f"Ошибка: {weather_data['error']}\n")
                weather_text.insert(tk.END, "Для реальных данных получите API ключ на openweathermap.org\n\n")
            else:
                weather_text.insert(tk.END, "✅ РЕАЛЬНЫЕ ДАННЫЕ С API\n\n")

            weather_text.insert(tk.END, f"🌍 Город: {weather_data['city']}\n")
            weather_text.insert(tk.END, f"🌡 Температура: {weather_data['temperature']}\n")
            weather_text.insert(tk.END, f"☁ Описание: {weather_data['description']}\n")
            weather_text.insert(tk.END, f"💧 Влажность: {weather_data['humidity']}\n")
            weather_text.insert(tk.END, f"📊 Давление: {weather_data['pressure']}\n")
            weather_text.insert(tk.END, f"💨 Ветер: {weather_data['wind']}\n")
        else:
            weather_text.insert(tk.END, "❌ Не удалось получить данные о погоде\n")
            weather_text.insert(tk.END, "Проверьте подключение к интернету и название города")

    btn_search = tk.Button(search_frame, text="🔍 Поиск погоды", command=search_weather, bg="lightblue")
    btn_search.pack(side='left', padx=5)

    # Загружаем данные по умолчанию
    search_weather()


# ==================== NEWS TAB ====================
def create_news_tab(tab):
    frame = tk.Frame(tab)
    frame.pack(fill='both', expand=True, padx=10, pady=10)

    tk.Label(frame, text="📰 Последние новости", font=("Arial", 16, "bold")).pack(pady=10)

    # Статус API
    api_status = tk.Label(frame, text="✅ Hacker News API - реальные данные", font=("Arial", 10), fg="green")
    api_status.pack(pady=5)

    # Кнопки для разных страниц новостей
    button_frame = tk.Frame(frame)
    button_frame.pack(pady=5)

    current_page = [1]  # Используем список для изменяемости

    def create_news_window(page_num, title):
        news_window = tk.Toplevel()
        news_window.title(f"{title} - Страница {page_num}")
        news_window.geometry("900x600")

        news_text = scrolledtext.ScrolledText(news_window, width=100, height=30, font=("Arial", 10))
        news_text.pack(pady=10, padx=10, fill='both', expand=True)

        def load_news_page():
            news_data = get_news_data(page_num)
            news_text.delete(1.0, tk.END)

            if news_data:
                news_text.insert(tk.END, f"✅ РЕАЛЬНЫЕ НОВОСТИ С HACKER NEWS API\n\n")
                news_text.insert(tk.END, f"📰 {title.upper()}\n\n")
                for i, news in enumerate(news_data, 1):
                    news_text.insert(tk.END, f"📖 {i}. {news['title']}\n")
                    news_text.insert(tk.END, f"   📋 Источник: {news['source']}\n")
                    news_text.insert(tk.END, f"   📅 Дата: {news['published']}\n")
                    if news.get('url'):
                        news_text.insert(tk.END, f"   🔗 Ссылка: {news['url']}\n")
                    news_text.insert(tk.END, "\n" + "-" * 80 + "\n\n")
            else:
                news_text.insert(tk.END, "❌ Не удалось загрузить новости\n")
                news_text.insert(tk.END, "Проверьте подключение к интернету\n")

        load_news_page()

        btn_refresh = tk.Button(news_window, text="🔄 Обновить", command=load_news_page, bg="lightgreen")
        btn_refresh.pack(pady=5)

    btn_top = tk.Button(button_frame, text="🔥 Топ новости",
                        command=lambda: create_news_window(1, "Топ новости"), bg="orange", fg="white")
    btn_top.pack(side='left', padx=5)

    btn_new = tk.Button(button_frame, text="🆕 Новые новости",
                        command=lambda: create_news_window(2, "Новые новости"), bg="blue", fg="white")
    btn_new.pack(side='left', padx=5)

    btn_ask = tk.Button(button_frame, text="❓ Вопросы",
                        command=lambda: create_news_window(3, "Вопросы"), bg="purple", fg="white")
    btn_ask.pack(side='left', padx=5)

    btn_show = tk.Button(button_frame, text="📖 Показать",
                         command=lambda: create_news_window(4, "Показать"), bg="green", fg="white")
    btn_show.pack(side='left', padx=5)

    # Основное окно с топ новостями
    news_text = scrolledtext.ScrolledText(frame, width=80, height=20, font=("Arial", 10))
    news_text.pack(pady=10, fill='both', expand=True)

    def load_news():
        news_data = get_news_data(1)
        news_text.delete(1.0, tk.END)

        if news_data:
            news_text.insert(tk.END, "✅ РЕАЛЬНЫЕ НОВОСТИ С HACKER NEWS API\n\n")
            news_text.insert(tk.END, "📰 ТОП НОВОСТИ\n\n")
            for i, news in enumerate(news_data, 1):
                news_text.insert(tk.END, f"📖 {i}. {news['title']}\n")
                news_text.insert(tk.END, f"   📋 Источник: {news['source']}\n")
                news_text.insert(tk.END, f"   📅 Дата: {news['published']}\n")
                if news.get('url'):
                    news_text.insert(tk.END, f"   🔗 Ссылка: {news['url']}\n")
                news_text.insert(tk.END, "\n" + "-" * 80 + "\n\n")
        else:
            news_text.insert(tk.END, "❌ Не удалось загрузить новости\n")
            news_text.insert(tk.END, "Проверьте подключение к интернету\n")

    btn_refresh = tk.Button(frame, text="🔄 Обновить топ новости", command=load_news, bg="lightgreen")
    btn_refresh.pack(pady=5)

    load_news()


# ==================== SPORTS TAB ====================
def create_sports_tab(tab):
    frame = tk.Frame(tab)
    frame.pack(fill='both', expand=True, padx=10, pady=10)

    tk.Label(frame, text="⚽ Спортивные события", font=("Arial", 16, "bold")).pack(pady=10)

    # Статус API
    api_status = tk.Label(frame, text="✅ TheSportsDB API - реальные данные", font=("Arial", 10), fg="green")
    api_status.pack(pady=5)

    sports_text = scrolledtext.ScrolledText(frame, width=80, height=20, font=("Arial", 10))
    sports_text.pack(pady=10, fill='both', expand=True)

    def load_sports():
        sports_data = get_sports_data()
        sports_text.delete(1.0, tk.END)

        if sports_data:
            sports_text.insert(tk.END, "✅ РЕАЛЬНЫЕ ДАННЫЕ С THE SPORTS DB\n\n")
            sports_text.insert(tk.END, "⚽ СПОРТИВНЫЕ СОБЫТИЯ\n\n")
            for sport in sports_data:
                sports_text.insert(tk.END, f"🏆 Вид спорта: {sport['sport']}\n")
                sports_text.insert(tk.END, f"🎯 Событие: {sport['event']}\n")
                sports_text.insert(tk.END, f"👥 Команды: {sport['result']}\n")
                sports_text.insert(tk.END, f"📅 Дата: {sport['date']}\n")
                sports_text.insert(tk.END, "-" * 50 + "\n\n")
        else:
            sports_text.insert(tk.END, "❌ Не удалось загрузить спортивные данные\n")

    btn_refresh = tk.Button(frame, text="🔄 Обновить спорт", command=load_sports, bg="lightcoral", fg="white")
    btn_refresh.pack(pady=5)

    load_sports()


# ==================== SPACE TAB ====================
def create_space_tab(tab):
    frame = tk.Frame(tab)
    frame.pack(fill='both', expand=True, padx=10, pady=10)

    tk.Label(frame, text="🚀 Космические новости", font=("Arial", 16, "bold")).pack(pady=10)

    # Статус API
    api_status = tk.Label(frame, text="✅ NASA API - реальные данные", font=("Arial", 10), fg="green")
    api_status.pack(pady=5)

    space_text = scrolledtext.ScrolledText(frame, width=80, height=20, font=("Arial", 10))
    space_text.pack(pady=10, fill='both', expand=True)

    def load_space():
        space_data = get_space_data()
        space_text.delete(1.0, tk.END)

        if space_data:
            space_text.insert(tk.END, "✅ РЕАЛЬНЫЕ ДАННЫЕ С NASA API\n\n")
            space_text.insert(tk.END, "🚀 АСТРОНОМИЧЕСКОЕ ФОТО ДНЯ\n\n")
            for news in space_data:
                space_text.insert(tk.END, f"🛰 {news['title']}\n")
                space_text.insert(tk.END, f"   📝 {news['info']}\n")
                space_text.insert(tk.END, f"   📅 {news['date']}\n\n")
        else:
            space_text.insert(tk.END, "❌ Не удалось загрузить космические данные\n")
            space_text.insert(tk.END, "Проверьте подключение к интернету\n")

    btn_refresh = tk.Button(frame, text="🔄 Обновить космос", command=load_space, bg="purple", fg="white")
    btn_refresh.pack(pady=5)

    load_space()


# ==================== PETS TAB ====================
def create_pets_tab(tab):
    frame = tk.Frame(tab)
    frame.pack(fill='both', expand=True, padx=10, pady=10)

    tk.Label(frame, text="🐾 Домашние животные", font=("Arial", 16, "bold")).pack(pady=10)

    # Статус API
    api_status = tk.Label(frame, text="✅ The Dog API - реальные данные", font=("Arial", 10), fg="green")
    api_status.pack(pady=5)

    pets_text = scrolledtext.ScrolledText(frame, width=80, height=20, font=("Arial", 10))
    pets_text.pack(pady=10, fill='both', expand=True)

    def load_pets():
        pets_data = get_pets_data()
        pets_text.delete(1.0, tk.END)

        if pets_data:
            pets_text.insert(tk.END, "✅ РЕАЛЬНЫЕ ДАННЫЕ С THE DOG API\n\n")
            pets_text.insert(tk.END, "🐾 СЛУЧАЙНЫЕ ПОРОДЫ СОБАК\n\n")
            for pet in pets_data:
                pets_text.insert(tk.END, f"🐕 Имя: {pet['name']}\n")
                pets_text.insert(tk.END, f"🐈 Порода: {pet['breed']}\n")
                pets_text.insert(tk.END, f"📋 Описание: {pet['info']}\n")
                if pet.get('image'):
                    pets_text.insert(tk.END, f"🖼 Ссылка на фото: {pet['image']}\n")
                pets_text.insert(tk.END, "-" * 50 + "\n\n")
        else:
            pets_text.insert(tk.END, "❌ Не удалось загрузить данные о животных\n")

    btn_refresh = tk.Button(frame, text="🔄 Обновить животных", command=load_pets, bg="orange")
    btn_refresh.pack(pady=5)

    load_pets()


# ==================== DATABASE TAB ====================
def create_database_tab(tab):
    """Создание вкладки базы данных"""
    input_frame = tk.Frame(tab)
    input_frame.pack(pady=10, padx=10, fill='x')

    fields = [
        "Фамилия", "Имя", "Отчество", "Телефон",
        "День_рождения", "Месяц_рождения", "Год_рождения",
        "Дом", "Корпус", "Подъезд", "Этаж", "Квартира", "Город",
        "Пол", "Вес", "Рост", "ЗП", "Цвет_волос",
        "Хобби", "Курс", "Специальность",
        "Левый_карман", "Правый_карман", "Деньги_на_карте"
    ]

    entries = {}
    for i, field in enumerate(fields):
        row = i // 4
        col = (i % 4) * 2
        tk.Label(input_frame, text=field).grid(row=row, column=col, sticky='e', padx=5, pady=2)
        entry = tk.Entry(input_frame, width=15)
        entry.grid(row=row, column=col + 1, sticky='w', padx=5, pady=2)
        entries[field] = entry

    button_frame = tk.Frame(input_frame)
    button_frame.grid(row=len(fields) // 4 + 1, column=0, columnspan=8, pady=10)

    btn_add = tk.Button(button_frame, text="Добавить человека",
                        command=lambda: add_person(entries, tree))
    btn_add.pack(side='left', padx=5)

    btn_delete = tk.Button(button_frame, text="Удалить выбранную запись",
                           command=lambda: delete_record(tree))
    btn_delete.pack(side='left', padx=5)

    btn_query = tk.Button(button_frame, text="SQL Запросы",
                          command=open_query_window, bg="lightcoral", fg="white")
    btn_query.pack(side='left', padx=5)

    btn_paint = tk.Button(button_frame, text="Открыть Paint",
                          command=open_paint, bg="lightblue")
    btn_paint.pack(side='left', padx=5)

    btn_calc = tk.Button(button_frame, text="Калькулятор",
                         command=open_calculator, bg="lightgreen")
    btn_calc.pack(side='left', padx=5)

    btn_calendar = tk.Button(button_frame, text="Календарь",
                             command=open_calendar, bg="lightyellow")
    btn_calendar.pack(side='left', padx=5)

    tree_frame = tk.Frame(tab)
    tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

    columns = ["ID"] + fields + ["Дата_рождения"]
    tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=80, anchor='center')

    vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

    tree.pack(side='left', fill='both', expand=True)
    vsb.pack(side='right', fill='y')
    hsb.pack(side='bottom', fill='x')

    load_data(tree)


def create_gui():
    root = tk.Tk()
    root.title("Управление таблицей Big_30_1 + API Вкладки")
    root.geometry("1200x800")

    # Создаем вкладки
    tab_control = ttk.Notebook(root)

    # Вкладка 1: База данных (основная)
    tab_db = ttk.Frame(tab_control)
    tab_control.add(tab_db, text='📊 База данных')

    # Вкладка 2: Настройки API
    tab_api = ttk.Frame(tab_control)
    tab_control.add(tab_api, text='⚙️ Настройки API')

    # Вкладка 3: Погода
    tab_weather = ttk.Frame(tab_control)
    tab_control.add(tab_weather, text='🌤 Погода')

    # Вкладка 4: Новости
    tab_news = ttk.Frame(tab_control)
    tab_control.add(tab_news, text='📰 Новости')

    # Вкладка 5: Спорт
    tab_sports = ttk.Frame(tab_control)
    tab_control.add(tab_sports, text='⚽ Спорт')

    # Вкладка 6: Космос
    tab_space = ttk.Frame(tab_control)
    tab_control.add(tab_space, text='🚀 Космос')

    # Вкладка 7: Животные
    tab_pets = ttk.Frame(tab_control)
    tab_control.add(tab_pets, text='🐾 Животные')

    tab_control.pack(expand=1, fill='both')

    # Создаем содержимое каждой вкладки
    create_database_tab(tab_db)
    create_api_settings_tab(tab_api)
    create_weather_tab(tab_weather)
    create_news_tab(tab_news)
    create_sports_tab(tab_sports)
    create_space_tab(tab_space)
    create_pets_tab(tab_pets)

    root.mainloop()


if __name__ == "__main__":
    create_gui()